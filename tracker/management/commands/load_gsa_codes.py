import openpyxl
from django.core.management.base import BaseCommand
from tracker.models import NAICSCode, PSCCode

class Command(BaseCommand):
    help = 'Load NAICS and PSC codes from GSA Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the GSA Excel file')

    def handle(self, *args, **options):
        file_path = options['file_path']
        self.stdout.write(f'Loading from {file_path}...')

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        
        naics_count = 0
        psc_count = 0

        for sheet_name in sheets:
            sheet = wb[sheet_name]
            self.stdout.write(f'Parsing sheet {sheet_name}...')
            
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                continue
                
            code_idx = -1
            desc_idx = -1
            
            if not headers: continue

            for idx, h in enumerate(headers):
                if not h: continue
                h_str = str(h).lower()
                if 'code' in h_str:
                    code_idx = idx
                if 'title' in h_str or 'description' in h_str or 'name' in h_str:
                    desc_idx = idx

            if code_idx == -1 or desc_idx == -1:
                code_idx = 0
                desc_idx = 1

            for row in rows:
                if len(row) <= max(code_idx, desc_idx): continue
                code_val = str(row[code_idx]).strip() if row[code_idx] else ''
                desc_val = str(row[desc_idx]).strip() if row[desc_idx] else ''
                
                if not code_val: continue
                
                if code_val.isdigit() and len(code_val) in [5, 6]:
                    NAICSCode.objects.update_or_create(code=code_val, defaults={'description': desc_val})
                    naics_count += 1
                elif len(code_val) == 4:
                    PSCCode.objects.update_or_create(code=code_val, defaults={'description': desc_val})
                    psc_count += 1

        self.stdout.write(self.style.SUCCESS(f'Successfully loaded {naics_count} NAICS codes and {psc_count} PSC codes.'))
